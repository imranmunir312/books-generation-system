from pathlib import Path

from openpyxl import load_workbook

from src.db import get_session
from src.models import Book
from src.notifications import send_email_notification


REQUIRED_COLUMNS = {"title", "notes_on_outline_before"}


def normalize_cell(value) -> str:
    return "" if value is None else str(value).strip()


def import_books_from_excel(file_path: str) -> None:
    workbook_path = Path(file_path)

    if not workbook_path.exists():
        raise RuntimeError(f"Excel file not found: {workbook_path}")

    workbook = load_workbook(workbook_path)
    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_cell(value) for value in header_row]
    header_indexes = {header: index for index, header in enumerate(headers) if header}
    missing_columns = REQUIRED_COLUMNS - set(header_indexes)

    if missing_columns:
        raise RuntimeError(
            "Excel file is missing required columns: "
            + ", ".join(sorted(missing_columns))
        )

    session = get_session()
    imported_count = 0
    skipped_count = 0

    try:
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            title = normalize_cell(row[header_indexes["title"]])
            notes = normalize_cell(row[header_indexes["notes_on_outline_before"]])

            if not title or not notes:
                skipped_count += 1
                continue

            book = Book(
                title=title,
                notes_on_outline_before=notes,
                status_outline_notes="no",
            )
            session.add(book)
            imported_count += 1

        session.commit()

        message = (
            f"Imported {imported_count} book row(s) from {workbook_path}. "
            f"Skipped {skipped_count} row(s)."
        )
        print(message)
        send_email_notification(
            subject="Book input import completed",
            body=message,
        )

    except Exception as error:
        session.rollback()
        send_email_notification(
            subject="Book input import failed",
            body=f"Import failed for {workbook_path}: {error}",
        )
        raise

    finally:
        session.close()


if __name__ == "__main__":
    import_books_from_excel("data/books_input.xlsx")
